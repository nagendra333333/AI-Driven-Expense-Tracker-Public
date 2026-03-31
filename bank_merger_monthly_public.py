"""
Monthly Bank Statement Merger  v3
===================================
✔ INPUT_FOLDER  = INPUT_FOLDER/  (sibling folder next to this script)
✔ OUTPUT_FOLDER = OUTPUT_FOLDER/ (sibling folder next to this script)
✔ Bank detection = reads file CONTENT, not filename
✔ Supports       = Canara · HDFC · KVB · Federal · Yes · Indian Bank
                   + any unknown bank via generic column-sniff fallback
✔ Transfer rule  = same-day same-amount debit↔credit across banks → deleted

USAGE
-----
1. Copy all 4 bank statement files into the SAME folder as this script.
2. Run:  python bank_merger_monthly.py
3. Open: merged_bank_statements.xlsx  (created in the same folder)

No filename patterns to edit. No paths to configure. Just drop and run.
"""

import os, re, csv, sys
import json
from dotenv import load_dotenv
from datetime import datetime
from pathlib import Path

import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ─── FOLDER: always the directory that contains this script ──────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent
INPUT_FOLDER = SCRIPT_DIR / "INPUT_FOLDER"
OUTPUT_FOLDER= SCRIPT_DIR / "OUTPUT_FOLDER"
# OUTPUT_FILE is computed dynamically in main() based on last transaction date

# ─── NVIDIA API CLIENT ───────────────────────────────────────────────────────
load_dotenv(SCRIPT_DIR / ".env")
nvidia_client = OpenAI(
    base_url = "https://integrate.api.nvidia.com/v1",
    api_key  = os.environ.get("NVIDIA_API_KEY", ""),  # Set in .env file
)

# Predefined categories the AI must choose from
# ─── INCOME CATEGORIES (from your Money Manager app) ─────────────────────────
INCOME_CATEGORIES = [
    "Salary",
    "Allowance",
    "Bonus",
    "Freelance",
    "Petty cash",
    "Dividend",
    "Stock capital",
    "Stock profit",
    "Bank interest credit",
    "Gold Bond interest",
    "Debt funds profit",
    "Debt Fund capital",
    "Lending Inflow",
    "Miscellaneous",
    "Other",
]

# Income keyword rules — format: ("Category|sub-category", [keywords])
# sub-category is optional; use "" or omit the pipe to leave it blank.
INCOME_KEYWORD_RULES = [
    ("Salary",               ["salary", "tata consultancy", "tcs", "infosys",
                               "wipro", "hcl", "cognizant", "accenture",
                               "payroll", "neft cr-citi", "neft cr-hdfc",
                               "neft cr-icici", "neft cr-axis", "credited by"]),
    ("Dividend",             ["dividend", "div payout"]),
    ("Bank interest credit", ["interest credit", "int.cr", "int cr",
                               "savings interest", "fd interest"]),
    ("Gold Bond interest",   ["gold bond", "sgb interest", "rbi bond"]),
    ("Debt funds profit",    ["debt fund", "liquid fund", "overnight fund"]),
    ("Stock profit",         ["stock profit", "equity profit", "capital gain"]),
    ("Stock capital",        ["stock capital", "share redemption"]),
    ("Lending Inflow",       ["loan repayment received", "lending inflow",
                               "repayment from", "emi received"]),
    ("Bonus",                ["bonus", "incentive", "performance pay"]),
    ("Freelance",            ["freelance", "consultancy fee", "project payment"]),
    ("Allowance",            ["allowance", "hra", "conveyance"]),
    ("Other",                ["cashback", "refund", "reversal", "reward",
                              "upi", "imps", "neft", "rtgs"]),   # generic credits / unknown
]


def apply_income_rules(desc: str) -> tuple:
    """Return (income_category, sub_category) for a credit transaction.
    Always returns "Other" when nothing matches — never "Miscellaneous".
    """
    d = desc.lower()
    for entry, keywords in INCOME_KEYWORD_RULES:
        if any(kw in d for kw in keywords):
            cat = entry.split("|")[0].strip() if "|" in entry else entry
            sub = entry.split("|")[1].strip() if "|" in entry else ""
            # Never return Miscellaneous for income — use Other
            if cat == "Miscellaneous":
                cat = "Other"
            return cat, sub
    return "Other", ""


# ─── EXPENSE CATEGORIES (from your Money Manager app) ─────────────────────────
# Exact categories from your Money Manager app (do not rename)
EXPENSE_CATEGORIES = [
    "Food",
    "Investment",
    "Fuel",
    "Dth + Ott + Net + Mobile",
    "Other",
    "Electricity",
    "Rent",
    "Ironing",
    "Gas",
    "water",
    "Spouse",
    "Baby",
    "Pregnancy",
    "Transportation",
    "Home things",
    "Movie",
    "Electronic Gadgets",
    "Bike maintenance",
    "Medicine",
    "Miscellaneous",
    "Drink",
    "Gift",
    "Mobile accessories",
    "Haircut",
    "Cloth",
    "Family",
    "Car maintenance",
    "Tour",
    "Marriage",
    "lending",
]


# ─── KEYWORD RULES (instant, no API) ─────────────────────────────────────────
# Each entry: (category_string, [keywords — any match wins, case-insensitive])
# Checked top-to-bottom; first match wins. Add your own patterns here freely.
KEYWORD_RULES = [
    # Groww / Mutual Fund — must come BEFORE the general Investment rule
    # so it matches first and gets the correct sub-category
    ("Investment|Mutual Fund",  ["groww", "nach groww", "nach  groww",
                                  "mutual fund", "kuvera", "paytm money",
                                  "hdfc mf", "sbi mf", "icici pru", "axis mf",
                                  "nippon mf", "mirae", "quant mf"]),
    ("Investment",              ["zerodha", "sip", "nse", "bse",
                                  "angel brok", "upstox", "smallcase"]),
    ("Food",                    ["zomato", "swiggy", "dunzo", "blinkit", "bigbasket",
                                  "tea stall", "chai", "tea corner", "ss tea",
                                  "jiomart", "restaurant", "hotel food", "canteen",
                                  "mess", "dominos", "pizza", "kfc", "mcdonalds",
                                  "burger", "cafe", "bakery", "biryani", "thalaiva",
                                  "food"]),
    ("Drink",                   ["tasmac", "beer", "wine", "alcohol", "whisky",
                                  "liquor", "bar ", "beverages"]),
    ("Fuel",                    ["bpcl", "hpcl", "hp petrol", "indian oil", "iocl",
                                  "petrol", "diesel", "fuel", "hp gas bunk",
                                  "reliance petro"]),
    ("Dth + Ott + Net + Mobile",["airtel", "jio", "bsnl", "netflix", "hotstar",
                                  "amazon prime", "spotify", "zee5", "sonyliv",
                                  "dth", "tatasky", "dish tv", "recharge",
                                  "broadband", "fibernet", "internet bill"]),
    ("Electricity",             ["tangedco", "bescom", "msedcl", "electricity",
                                  "eb bill", "power bill", "torrent power",
                                  "cesc", "tneb"]),
    ("Gas",                     ["indane", "hp gas", "bharat gas", "lpg", "gas cyl",
                                  "gas book"]),
    ("water",                   ["water can", "water bill", "drinking water",
                                  "bwssb", "cmwssb"]),
    ("Rent",                    ["rent", "house rent", "room rent", "rental"]),
    ("Ironing",                 ["ironing", "laundry", "dhobi", "clothes press"]),
    ("Medicine",                ["pharmacy", "medplus", "apollo pharmacy",
                                  "netmeds", "1mg", "hospital", "clinic",
                                  "doctor", "medical", "medicine", "health"]),
    ("Movie",                   ["bookmyshow", "pvr", "inox", "ags cinema",
                                  "agscinema", "cinepolis", "movie", "theatre",
                                  "multiplex"]),
    ("Transportation",          ["ola", "uber", "rapido", "redbus",
                                  "irctc", "metro", "auto pay", "cab",
                                  "bus ticket", "transport"]),
    ("Tour",                    ["makemytrip", "goibibo", "yatra", "cleartrip",
                                  "flight", "airline", "indigo", "air india",
                                  "hotel booking", "oyo", "treebo", "tour"]),
    ("Electronic Gadgets",      ["reliance digital", "croma", "vijay sales",
                                  "flipkart electron", "laptop", "tablet",
                                  "earphones", "headphone", "smartwatch", "gadget"]),
    ("Bike maintenance",        ["bike service", "two wheeler", "bike repair",
                                  "tyre", "puncture", "honda service",
                                  "bajaj service", "hero service", "tvs service"]),
    ("Car maintenance",         ["car service", "four wheeler", "car repair",
                                  "maruti service", "hyundai service",
                                  "car wash", "automobile"]),
    ("Mobile accessories",      ["mobile cover", "screen guard", "charger",
                                  "mobile accessories", "earbuds", "cable"]),
    ("Haircut",                 ["salon", "haircut", "barber", "beauty parlour",
                                  "hair cut", "trimming"]),
    ("Cloth",                   ["myntra", "meesho", "ajio", "lifestyle",
                                  "pantaloons", "westside", "cloth", "dress",
                                  "saree", "shirt", "fashion", "textile"]),
    ("Gift",                    ["gift", "birthday", "anniversary", "present",
                                  "amazon gift"]),
    ("Home things",             ["ikea", "pepperfry", "urban ladder", "nilkamal",
                                  "furniture", "household", "utensils",
                                  "kitchen", "home decor", "cleaning"]),
    ("Family",                  ["spouse", "wife", "mother", "father", "parents",
                                  "brother", "sister", "amma", "appa"]),
    ("Marriage",                ["wedding", "marriage", "reception", "muhurtham"]),
    ("lending",                 ["lending", "loan given", "borrowed by"]),
    ("Baby",                    ["baby", "diaper", "infant", "toddler",
                                  "pampers", "huggies"]),
    ("Pregnancy",               ["pregnancy", "maternity", "gynaec", "prenatal",
                                  "antenatal"]),
    # ── Common UPI merchant patterns ─────────────────────────────────────────
    ("Food",                    ["grocery", "grocer", "kirana", "provision",
                                  "supermarket", "mani store", "stores",
                                  "meat", "chicken", "mutton", "vegetable",
                                  "fruits", "ration"]),
    ("Transportation",          ["cumta", "kmrl", "dmrc", "bmtc", "ksrtc",
                                  "state transport", "bus pass", "fastag",
                                  "toll", "parking"]),
    ("Other",                   ["tata consultancy", "infosys", "wipro",
                                  "hcl tech", "cognizant", "salary",
                                  "payroll", "neft cr-citi", "neft cr-hdfc",
                                  "annual fee", "chrg", "charges",
                                  "service charge", "gst debit",
                                  "debit card fee", "card fee", "joining fee",
                                  "amazon", "flipkart", "snapdeal",
                                  "cashback", "refund", "reversal"]),
    ("Other",                   ["upi-dr", "upiout", "upi out", "upi/dr",
                                  "paid to", "send money"]),
]

# Descriptions that match these keywords are sent to API (truly ambiguous)
AMBIGUOUS_KEYWORDS = ["upi", "neft", "imps", "payment", "transfer", "paid",
                      "cr-", "dr-", "by order", "mob"]


def apply_rules(desc: str) -> tuple | None:
    """
    Try to categorize using keyword rules.
    Returns (category, sub_category) if matched, None if unknown.
    Category entries may be "Cat|Sub" format to also set sub-category.
    """
    d = desc.lower()
    for category, keywords in KEYWORD_RULES:
        if any(kw in d for kw in keywords):
            if "|" in category:
                cat, sub = category.split("|", 1)
                return cat.strip(), sub.strip()
            return category, ""
    return None   # couldn't decide — send to API


def clean_desc(d: str) -> str:
    """Strip long UPI hash codes and timestamps for cleaner AI input."""
    d = str(d)
    d = re.sub(r'/[A-Za-z0-9]{20,}', '', d)
    d = re.sub(r'\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}$', '', d.strip())
    d = re.sub(r'[/\s]+', ' ', d).strip()
    return d[:100]


def _api_batch(args: tuple) -> tuple[int, list[str]]:
    """
    Worker function — called in a thread.
    args = (batch_index, [(original_index, description), ...])
    Returns (batch_index, [category, ...]) in original index order.
    """
    import time
    b_idx, pairs = args
    indices  = [p[0] for p in pairs]
    cleaned  = [clean_desc(p[1]) for p in pairs]
    cats_str = ", ".join(EXPENSE_CATEGORIES)
    numbered = "\n".join(f"{i}: {d}" for i, d in enumerate(cleaned))

    prompt = (
        "Indian bank transaction categorizer.\n"
        f"Pick ONE category from: {cats_str}\n\n"
        "Groww/SIP/Zerodha/NACH -> Investment | "
        "Zomato/Swiggy/food -> Food | "
        "Tea/chai -> Food | TASMAC/beer -> Drink | "
        "Petrol/BPCL/fuel -> Fuel | "
        "Airtel/Jio/Netflix/DTH -> Dth + Ott + Net + Mobile | "
        "EB/electricity -> Electricity | Gas/LPG -> Gas | "
        "Water -> water | Rent -> Rent | Ironing -> Ironing | "
        "Spouse/partner name -> Spouse | Baby/diaper -> Baby | "
        "Pregnancy -> Pregnancy | Ola/Uber/IRCTC -> Transportation | "
        "Home items -> Home things | Cinema/AGS -> Movie | "
        "Electronics/gadgets -> Electronic Gadgets | "
        "Bike service -> Bike maintenance | "
        "Pharmacy/hospital -> Medicine | Gift -> Gift | "
        "Mobile cover -> Mobile accessories | "
        "Salon/haircut -> Haircut | Cloth/Myntra -> Cloth | "
        "Family transfer -> Family | Car service -> Car maintenance | "
        "MakeMyTrip/flight/hotel -> Tour | Wedding -> Marriage | "
        "Loan given -> lending | Unclear person UPI -> R | "
        "Bank charges -> Other | Salary -> Other | "
        "Amazon/Flipkart -> Other | Unclear -> Miscellaneous\n\n"
        f"Transactions:\n{numbered}\n\n"
        f"Return ONLY a JSON array of exactly {len(pairs)} strings.\n"
        'Example: ["Food","Investment","Transportation"]\n'
        "No explanation. No markdown. Just the JSON array."
    )

    valid = set(EXPENSE_CATEGORIES)
    for attempt in range(3):
        wait = 3 * (2 ** attempt)   # 3s, 6s, 12s
        try:
            completion = nvidia_client.chat.completions.create(
                model       = "openai/gpt-oss-120b",
                messages    = [{"role": "user", "content": prompt}],
                temperature = 0.1,
                top_p       = 1,
                max_tokens  = 200,
                stream      = True,
            )
            raw = ""
            for chunk_part in completion:
                if not chunk_part.choices:
                    continue
                delta = chunk_part.choices[0].delta.content
                if delta is not None:
                    raw += delta

            if not raw.strip():
                time.sleep(wait)
                continue

            match = re.search(r'\[.*?\]', raw, re.DOTALL)
            if match:
                cats = json.loads(match.group())
                out  = [c.strip() if c.strip() in valid else "Other"
                        for c in cats[:len(pairs)]]
                # Pad if AI returned fewer items than expected
                while len(out) < len(pairs):
                    out.append("Other")
                return (b_idx, out)

        except Exception:
            time.sleep(wait)

    # All attempts failed
    return (b_idx, ["Other"] * len(pairs))


def categorize_transactions(descriptions: list[str]) -> list[str]:
    """
    Hybrid + Parallel categorization:
      Phase 1 — Keyword rules  : instant, handles ~80% of transactions locally
      Phase 2 — Parallel API   : remaining unknowns sent in parallel threads
                                  (max 5 concurrent calls, batch size 5)
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time

    BATCH      = 5    # per API call
    MAX_WORKERS = 5   # simultaneous API threads

    results = [None] * len(descriptions)

    # ── Phase 1: keyword rules ────────────────────────────────────────────────
    # results_sub tracks sub-category from rules (e.g. "Mutual Fund" for Groww)
    results_sub     = [""] * len(descriptions)
    unknown_indices = []
    rule_hits       = 0
    for i, desc in enumerate(descriptions):
        match = apply_rules(desc)
        if match:
            results[i], results_sub[i] = match
            rule_hits += 1
        else:
            unknown_indices.append(i)

    print(f"    Rules matched : {rule_hits}/{len(descriptions)} instantly")
    print(f"    Sending to API: {len(unknown_indices)} unknowns", flush=True)

    if not unknown_indices:
        return [r or "Other" for r in results], results_sub

    # ── Phase 2: parallel API for unknowns ───────────────────────────────────
    # Build batches of (original_index, description)
    pairs  = [(i, descriptions[i]) for i in unknown_indices]
    batches = [
        (b, pairs[b * BATCH : (b + 1) * BATCH])
        for b in range((len(pairs) + BATCH - 1) // BATCH)
    ]
    total = len(batches)
    done  = 0

    print(f"    Running {total} API batches across {MAX_WORKERS} parallel threads...")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_api_batch, batch): batch[0] for batch in batches}
        for future in as_completed(futures):
            b_idx, cats = future.result()
            batch_pairs = batches[b_idx][1]
            for k, (orig_i, _) in enumerate(batch_pairs):
                results[orig_i] = cats[k] if k < len(cats) else "Other"
            done += 1
            print(f"    ✓ {done}/{total} batches done", end="\r", flush=True)

    print(f"    ✓ All {total} API batches complete          ")

    return [r or "Other" for r in results], results_sub









# ─── TRADEBOOK FILENAME PATTERN ─────────────────────────────────────────────
# Any file whose name contains "tradebook" (case-insensitive) is treated as a
# Zerodha tradebook, not a bank statement. It gets parsed separately.
TRADEBOOK_PATTERN = "tradebook"


# ─── BANK NAME MAP ───────────────────────────────────────────────────────────
# Key   = word that must appear in the filename (case-insensitive)
# Value = (display label,  parser key)
# Add a new row here whenever you add a new bank.
BANK_NAME_MAP = {
    "canara":  ("Canara Bank",  "Canara"),
    "hdfc":    ("HDFC Bank",    "HDFC"),
    "kvb":     ("KVB Bank",     "KVB"),
    "federal": ("Federal Bank", "Federal"),
    "yes":     ("Yes Bank",     "Yes"),
    "indian":  ("Indian Bank",  "Indian"),
    "idbi":    ("IDBI Bank",    "Indian"),
    "sbi":     ("SBI",          "Unknown"),
    "axis":    ("Axis Bank",    "Unknown"),
    "icici":   ("ICICI Bank",   "Unknown"),
    "kotak":   ("Kotak Bank",   "Unknown"),
}


def identify_bank(path: Path) -> tuple:
    """
    Identify bank from the filename stem (case-insensitive partial match).
    Returns (display_label, parser_key).
    e.g.  CANARA.xlsx  →  ('Canara Bank', 'Canara')
          HDFC.xls     →  ('HDFC Bank',   'HDFC')
    """
    stem = path.stem.lower()
    for keyword, (display, parser_key) in BANK_NAME_MAP.items():
        if keyword in stem:
            return display, parser_key
    return "Unknown Bank", "Unknown"



def to_float(s):
    try:
        return float(str(s).strip().replace(",", ""))
    except Exception:
        return None


def parse_date_str(s):
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%y",
                "%Y-%m-%d", "%d %b %Y", "%d-%b-%Y", "%d/%b/%Y"):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except Exception:
            pass
    return pd.NaT


def _read_lines(path: Path):
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(path, encoding=enc) as f:
                return f.readlines()
        except UnicodeDecodeError:
            continue
    return []


# ─── PARSERS ─────────────────────────────────────────────────────────────────

def _strip_eq(v: str) -> str:
    v = v.strip()
    if v.startswith('="') and v.endswith('"'):
        v = v[2:-1]
    elif v.startswith("="):
        v = v[1:]
    return v.strip('"').strip()


def parse_canara(path: Path) -> pd.DataFrame:
    lines = _read_lines(path)
    hdr_idx = next((i for i, l in enumerate(lines) if l.startswith("Txn Date")), None)
    if hdr_idx is None:
        return pd.DataFrame()
    rows = []
    reader = csv.reader(lines[hdr_idx:])
    next(reader)
    for row in reader:
        if not row or len(row) < 7:
            continue
        date = _strip_eq(row[0]).split()[0]
        desc = _strip_eq(row[3])
        deb  = _strip_eq(row[5])
        cred = _strip_eq(row[6])
        if not date or date == "Txn Date":
            continue
        rows.append({"Date": date, "Description": desc, "Debit": deb, "Credit": cred})
    return pd.DataFrame(rows)


def parse_hdfc(path: Path) -> pd.DataFrame:
    eng = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    raw = pd.read_excel(path, engine=eng, header=None, dtype=str)
    hdr_idx = None
    for i, row in raw.iterrows():
        if str(row[0]).strip() in ("Date", "********") and "Narration" in str(row[1]):
            hdr_idx = i
            break
    if hdr_idx is None:
        return pd.DataFrame()
    data = raw.iloc[hdr_idx + 2:].reset_index(drop=True)
    rows = []
    for _, row in data.iterrows():
        date = str(row[0]).strip()
        narr = str(row[1]).strip()
        wa   = str(row[4]).strip()
        da   = str(row[5]).strip()
        if date in ("nan", "", "Date", "********") or narr == "nan":
            continue
        if not re.search(r"\d{2}[/\-]\d{2}", date):
            continue
        deb  = "" if wa in ("nan", "") else wa
        cred = "" if da in ("nan", "") else da
        rows.append({"Date": date, "Description": narr, "Debit": deb, "Credit": cred})
    return pd.DataFrame(rows)


def parse_yes(path: Path) -> pd.DataFrame:
    lines = _read_lines(path)
    hdr_idx = next((i for i, l in enumerate(lines) if l.startswith("Transaction Date")), None)
    if hdr_idx is None:
        return pd.DataFrame()
    rows = []
    reader = csv.reader(lines[hdr_idx:])
    next(reader)
    for row in reader:
        if not row or len(row) < 7:
            continue
        date = row[0].strip().split()[0]
        desc = row[4].strip()
        deb  = row[5].strip().strip('"')
        cred = row[6].strip().strip('"')
        if not date or date == "Transaction Date":
            continue
        rows.append({"Date": date, "Description": desc, "Debit": deb, "Credit": cred})
    return pd.DataFrame(rows)


def parse_indian(path: Path) -> pd.DataFrame:
    lines = _read_lines(path)
    hdr_idx = next((i for i, l in enumerate(lines) if "Sl. No." in l), None)
    if hdr_idx is None:
        return pd.DataFrame()
    rows = []
    reader = csv.reader(lines[hdr_idx:])
    next(reader)
    for row in reader:
        if not row or len(row) < 9:
            continue
        if not str(row[0]).strip().isdigit():
            continue
        deb  = str(row[7]).strip().strip('"')
        cred = str(row[8]).strip().strip('"')
        rows.append({"Date": str(row[1]).strip(), "Description": str(row[2]).strip(),
                     "Debit": deb, "Credit": cred})
    return pd.DataFrame(rows)


def _column_sniff_parse(path: Path) -> pd.DataFrame:
    """
    Generic parser used for KVB, Federal, and any unrecognised bank.
    Scans for the header row by looking for date + debit/credit keywords,
    then auto-maps column indices by header name.
    Works for most standard Indian bank CSV exports.
    """
    SKIP_ROWS = {"opening balance", "closing balance", "total", "b/f", "c/f",
                 "brought forward", "carried forward", "balance b/f"}

    # Try as CSV first
    lines = _read_lines(path)

    hdr_idx = None
    if lines:
        for i, line in enumerate(lines):
            ll = line.lower()
            if "date" in ll and ("debit" in ll or "withdrawal" in ll or "credit" in ll or "deposit" in ll):
                hdr_idx = i
                break

    if hdr_idx is not None:
        reader  = csv.reader(lines[hdr_idx:])
        headers = [h.strip().lower() for h in next(reader)]

        def col(*names):
            for n in names:
                for i, h in enumerate(headers):
                    if n in h:
                        return i
            return None

        di  = col("txn date", "transaction date", "trans date", "value date", "date")
        dsi = col("narration", "description", "particulars", "transaction details", "remarks", "details")
        dbi = col("withdrawal", "debit", " dr", "dr amount")
        cri = col("deposit", "credit", " cr", "cr amount")

        rows = []
        for row in reader:
            if not row:
                continue
            get  = lambda idx: row[idx].strip() if idx is not None and idx < len(row) else ""
            date = get(di).split()[0] if get(di) else ""
            desc = get(dsi)
            deb  = get(dbi)
            cred = get(cri)
            if not date:
                continue
            if any(kw in desc.lower() for kw in SKIP_ROWS):
                continue
            rows.append({"Date": date, "Description": desc, "Debit": deb, "Credit": cred})
        if rows:
            return pd.DataFrame(rows)

    # Fallback: try Excel (XLS / XLSX)
    if path.suffix.lower() in (".xls", ".xlsx"):
        try:
            eng = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
            raw = pd.read_excel(path, engine=eng, header=None, dtype=str)
            hdr_idx = None
            for i, row in raw.iterrows():
                vals = " ".join(str(v).lower() for v in row if pd.notna(v))
                if "date" in vals and ("debit" in vals or "credit" in vals):
                    hdr_idx = i
                    break
            if hdr_idx is not None:
                raw.columns = range(raw.shape[1])
                headers = [str(raw.iloc[hdr_idx, c]).strip().lower() for c in range(raw.shape[1])]
                data = raw.iloc[hdr_idx + 1:].reset_index(drop=True)

                def col(*names):
                    for n in names:
                        for i, h in enumerate(headers):
                            if n in h:
                                return i
                    return None

                di  = col("txn date", "transaction date", "date")
                dsi = col("narration", "description", "particulars")
                dbi = col("withdrawal", "debit", "dr")
                cri = col("deposit", "credit", "cr")

                rows = []
                for _, row in data.iterrows():
                    get  = lambda idx: str(row[idx]).strip() if idx is not None else ""
                    date = get(di).split()[0] if get(di) else ""
                    desc = get(dsi)
                    deb  = get(dbi) if get(dbi) not in ("nan", "") else ""
                    cred = get(cri) if get(cri) not in ("nan", "") else ""
                    if not date or date in ("nan", ""):
                        continue
                    if not re.search(r"\d{2}[/\-]\d{2}", date):
                        continue
                    rows.append({"Date": date, "Description": desc, "Debit": deb, "Credit": cred})
                return pd.DataFrame(rows)
        except Exception as e:
            print(f"    [WARN] Excel parse failed for {path.name}: {e}")

    print(f"    [WARN] Could not parse {path.name} — skipping")
    return pd.DataFrame()


# Map bank label → parser
PARSERS = {
    "Canara":  parse_canara,
    "HDFC":    parse_hdfc,
    "Yes":     parse_yes,
    "Indian":  parse_indian,
    "KVB":     _column_sniff_parse,   # KVB uses generic sniff
    "Federal": _column_sniff_parse,   # Federal uses generic sniff
    "Unknown": _column_sniff_parse,
}


# ─── BANK-TO-BANK TRANSFER REMOVAL ───────────────────────────────────────────

def remove_bank_transfers(merged: pd.DataFrame) -> pd.DataFrame:
    """
    Rule: same calendar date + same amount appearing as Debit in one bank
    and Credit in a DIFFERENT bank → both rows are a transfer → delete both.
    1-to-1 matching prevents over-deletion of repeated same-amount transactions.
    """
    merged = merged.reset_index(drop=True)
    merged["_rid"] = merged.index

    debit_rows  = merged[(merged["_amt_deb"].notna())  & (merged["_amt_deb"]  > 0)][
        ["_rid", "_norm_date", "_amt_deb", "_bank", "Description"]].copy()
    credit_rows = merged[(merged["_amt_cred"].notna()) & (merged["_amt_cred"] > 0)][
        ["_rid", "_norm_date", "_amt_cred", "_bank", "Description"]].copy()

    pairs = debit_rows.merge(
        credit_rows,
        left_on  = ["_norm_date", "_amt_deb"],
        right_on = ["_norm_date", "_amt_cred"],
        suffixes = ("_d", "_c"),
    )
    pairs = pairs[pairs["_bank_d"] != pairs["_bank_c"]].reset_index(drop=True)

    drop_ids  = set()
    used_drid = set()
    used_crid = set()
    found     = []

    for _, p in pairs.iterrows():
        d_id = p["_rid_d"]; c_id = p["_rid_c"]
        if d_id in used_drid or c_id in used_crid:
            continue
        used_drid.add(d_id); used_crid.add(c_id)
        drop_ids.add(d_id);  drop_ids.add(c_id)
        found.append(p)

    if found:
        print(f"\n{'─'*70}")
        print("  Bank-to-Bank Transfers Detected & Removed:")
        print(f"{'─'*70}")
        for p in found:
            print(f"  {p['_norm_date']}  ₹{p['_amt_deb']:>10,.2f}"
                  f"   DEBIT [{p['_bank_d']}] → CREDIT [{p['_bank_c']}]")
            print(f"    ↳ {str(p['Description_d'])[:60]}")
            print(f"    ↳ {str(p['Description_c'])[:60]}")
        n = len(drop_ids) // 2
        print(f"\n  → {len(drop_ids)} rows removed ({n} transfer pair{'s' if n != 1 else ''})")
        print(f"{'─'*70}")
    else:
        print("\n  No bank-to-bank transfers detected.")

    transfer_df = merged[merged["_rid"].isin(drop_ids)].copy()
    clean_df    = merged[~merged["_rid"].isin(drop_ids)].copy()
    return clean_df, transfer_df


# ─── TRADEBOOK PARSER ────────────────────────────────────────────────────────

def parse_tradebook(path: Path) -> pd.DataFrame:
    """
    Parse a Zerodha tradebook CSV.
    Aggregates by (symbol, trade_date) → total quantity & total amount spent.
    Returns rows ready to merge with bank transactions:
      Date | Description | Debit | Credit | Bank | Expenses Category | Expenses sub-category
    """
    try:
        df = pd.read_csv(str(path), dtype=str)
    except Exception as e:
        print(f"  ✗ Tradebook read error: {e}")
        return pd.DataFrame()

    required = {"symbol", "trade_date", "quantity", "price"}
    if not required.issubset(set(df.columns)):
        print(f"  ✗ Tradebook missing columns. Found: {df.columns.tolist()}")
        return pd.DataFrame()

    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0)
    df["price"]    = pd.to_numeric(df["price"],    errors="coerce").fillna(0)
    df["amount"]   = df["quantity"] * df["price"]

    # Aggregate: same symbol bought on same date → sum qty & amount
    agg = (
        df.groupby(["symbol", "trade_date"], sort=True)
          .agg(total_qty=("quantity", "sum"),
               total_amount=("amount",   "sum"))
          .reset_index()
    )

    rows = []
    for _, r in agg.iterrows():
        rows.append({
            "Date":                  r["trade_date"],
            "Description":           f"Zerodha: {r['symbol']} x{int(r['total_qty'])} shares",
            "Debit":                 f"{r['total_amount']:.2f}",
            "Credit":                "",
            "_bank":                 "Zerodha",
            "Expenses Category":     "Investment",
            "Expenses sub-category": r["symbol"],
            "_amt_deb":              r["total_amount"],
            "_amt_cred":             None,
        })

    out = pd.DataFrame(rows)
    print(f"  ✓ {'Zerodha':<14}  {len(out):>4} rows (aggregated from {len(df)} trades)  ←  {path.name}")
    return out


# ─── EXCEL WRITER ─────────────────────────────────────────────────────────────

def load_category_history(output_folder: Path, current_file: str = "") -> dict:
    """
    Scan all Full_statement_*.xlsx files in OUTPUT_FOLDER.
    Build a lookup: cleaned_description → (Expenses Category, Expenses sub-category).
    Skip the file currently being written (current_file).
    Returns dict keyed by lowercased, stripped description.
    """
    history = {}
    pattern = "Full_statement_*.xlsx"
    files   = sorted(output_folder.glob(pattern))
    loaded  = 0
    for fp in files:
        if fp.name == current_file:
            continue
        try:
            df = pd.read_excel(str(fp), sheet_name="Transactions", dtype=str)
            for _, row in df.iterrows():
                desc = str(row.get("Description", "")).strip().lower()
                cat  = str(row.get("Expenses Category", "")).strip()
                sub  = str(row.get("Expenses sub-category", "")).strip()
                if desc and cat and cat not in ("", "nan", "Other", "Miscellaneous"):
                    if desc not in history:           # first occurrence wins
                        history[desc] = (cat, sub)
            loaded += 1
        except Exception as e:
            print(f"    [history] Could not read {fp.name}: {e}")
    if loaded:
        print(f"  History loaded: {len(history)} known descriptions from {loaded} previous file(s)")
    return history


def match_history(desc: str, history: dict) -> tuple | None:
    """
    Try exact match, then partial match (description contains a known key).
    Returns (category, sub_category) or None.
    """
    key = desc.strip().lower()
    if key in history:
        return history[key]
    # Partial: check if any known description is a substring of current
    for known, cats in history.items():
        if len(known) > 10 and known in key:
            return cats
    return None


def _resolve_output_path(output_path: Path) -> Path:
    """
    Windows locks xlsx files that are open in Excel.
    Strategy:
      1. Try to force-close Excel on Windows via taskkill (silent).
      2. Wait briefly and retry.
      3. If still locked → save to a timestamped fallback name instead.
    Returns the path that was actually used.
    """
    def is_locked(p: Path) -> bool:
        if not p.exists():
            return False
        try:
            p.rename(p)   # zero-cost probe — fails if locked
            return False
        except PermissionError:
            return True

    if not is_locked(output_path):
        return output_path

    print()
    print("  ⚠  Output file is open in Excel — attempting to close Excel…")

    # Try closing Excel on Windows
    if sys.platform == "win32":
        import subprocess, time
        subprocess.run(
            ["taskkill", "/F", "/IM", "EXCEL.EXE"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )
        time.sleep(2)   # give Windows time to release the file handle

    if not is_locked(output_path):
        print("  ✓  Excel closed successfully — saving to original filename.")
        return output_path

    # Still locked → use a timestamped fallback name
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem     = output_path.stem   # e.g. "Full_statement_2026-03-23"
    fallback = output_path.parent / f"{stem}_{ts}.xlsx"
    print(f"  ℹ  File still locked — saving to fallback:")
    print(f"     {fallback.name}")
    return fallback


def write_excel(df: pd.DataFrame, transfer_df: pd.DataFrame, output_path: Path):
    output_path = _resolve_output_path(output_path)

    COLS = ["S.No", "Date", "Description", "Debit", "Credit", "Bank",
            "Expenses Category", "Expenses sub-category",
            "Income Category",   "Income sub-category"]

    # Ensure all columns exist and are clean (no 'nan' string values)
    for col in COLS:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).str.strip()
        df[col] = df[col].replace({"nan": "", "NaN": "", "None": ""})

    df[COLS].to_excel(str(output_path), sheet_name="Transactions", index=False)

    wb = load_workbook(str(output_path))

    # ── Hidden Lists sheet for dropdown validation ──────────────────────────
    ws_lists = wb.create_sheet("_Lists")
    ws_lists.sheet_state = "hidden"

    exp_cats = EXPENSE_CATEGORIES
    inc_cats = INCOME_CATEGORIES
    exp_subs = sorted({str(v) for v in df["Expenses sub-category"].dropna() if str(v).strip()})
    inc_subs = sorted({str(v) for v in df["Income sub-category"].dropna()   if str(v).strip()})
    # Pad all lists to same length
    max_len = max(len(exp_cats), len(inc_cats), len(exp_subs) or 1, len(inc_subs) or 1)
    for r in range(max_len):
        ws_lists.cell(r+1, 1, exp_cats[r] if r < len(exp_cats) else "")
        ws_lists.cell(r+1, 2, inc_cats[r] if r < len(inc_cats) else "")
        ws_lists.cell(r+1, 3, exp_subs[r] if r < len(exp_subs) else "")
        ws_lists.cell(r+1, 4, inc_subs[r] if r < len(inc_subs) else "")

    # ── Main Transactions sheet ─────────────────────────────────────────────
    ws = wb["Transactions"]
    n_data = len(df)

    HDR_FILL  = PatternFill("solid", fgColor="1B2A4A")
    HDR_FONT  = Font(name="Calibri", bold=True, color="E8F0FE", size=11)
    ALT_BG    = PatternFill("solid", fgColor="F4F6F9")
    WH_BG     = PatternFill("solid", fgColor="FFFFFF")
    CRED_BG   = PatternFill("solid", fgColor="E8F8EE")
    DEB_BG    = PatternFill("solid", fgColor="FEF0EF")
    INC_CAT   = PatternFill("solid", fgColor="ECFDF5")
    INC_SUB   = PatternFill("solid", fgColor="D1FAE5")
    THIN      = Side(style="thin", color="C8D0DE")
    BDR       = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    NUM_FMT   = '#,##0.00'

    COL_WIDTHS = {"A":7,"B":13,"C":54,"D":14,"E":14,"F":16,"G":22,"H":18,"I":22,"J":18}
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 23

    CAT_COLORS = {
        "Food":("FFF8E7","92400E"),"Investment":("D1FAE5","064E3B"),
        "Fuel":("FEF3C7","78350F"),"Dth + Ott + Net + Mobile":("EDE9FE","4C1D95"),
        "Electricity":("FEF9C3","713F12"),"Rent":("FFE4E6","9F1239"),
        "Ironing":("F0FDF4","166534"),"Gas":("FFF3E0","B45309"),
        "water":("E0F2FE","0369A1"),"Spouse":("FCE7F3","9D174D"),
        "Baby":("FDF4FF","7E22CE"),"Pregnancy":("FDF2F8","9D174D"),
        "Transportation":("EFF6FF","1E40AF"),"Home things":("F5F3FF","5B21B6"),
        "Movie":("FFF1F2","9F1239"),"Electronic Gadgets":("F0F9FF","0C4A6E"),
        "Bike maintenance":("FEF3C7","92400E"),"Medicine":("FCE7F3","9D174D"),
        "Miscellaneous":("F9FAFB","374151"),"Drink":("FEF2F2","991B1B"),
        "Gift":("FDF4FF","6B21A8"),"Mobile accessories":("ECFDF5","065F46"),
"Haircut":("F0FDF4","166534"),
        "Cloth":("FDF4FF","7E22CE"),"Family":("DCFCE7","14532D"),
        "Car maintenance":("FEF3C7","B45309"),"Tour":("E0F2FE","0369A1"),
        "Marriage":("FDF2F8","BE185D"),"lending":("FEE2E2","991B1B"),
"Other":("F9FAFB","374151"),
    }

    INC_COLORS = {
        "Salary":("DCFCE7","14532D"),"Bonus":("D1FAE5","065F46"),
        "Freelance":("E0F2FE","0369A1"),"Dividend":("FEF9C3","713F12"),
        "Stock profit":("DCFCE7","064E3B"),"Stock capital":("D1FAE5","064E3B"),
        "Bank interest credit":("FFF8E7","92400E"),
        "Gold Bond interest":("FEF9C3","78350F"),
        "Lending Inflow":("FEE2E2","991B1B"),"Allowance":("EDE9FE","4C1D95"),
        "Miscellaneous":("F9FAFB","374151"),"Other":("F9FAFB","374151"),
    }

    # Header row
    for cell in ws[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BDR

    # Data rows
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = ALT_BG if i % 2 == 0 else WH_BG
        ws.row_dimensions[i].height = 15
        for cell in row:
            cell.border = BDR
            c = cell.column
            if c == 1:   # S.No
                cell.fill = bg
                cell.font = Font(name="Calibri", size=9, color="6B7280")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 2:   # Date
                cell.fill = bg
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 3:   # Description
                cell.fill = bg
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c == 4:   # Debit — proper number format
                cell.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    v = float(str(cell.value).replace(",","")) if cell.value and str(cell.value).strip() not in ("","—","nan") else None
                except (ValueError, TypeError):
                    v = None
                if v:
                    cell.value         = v
                    cell.number_format = NUM_FMT
                    cell.fill = DEB_BG
                    cell.font = Font(name="Calibri", size=10, bold=True, color="C62828")
                else:
                    cell.value = None; cell.fill = bg
                    cell.font  = Font(name="Calibri", size=10, color="CCCCCC")
            elif c == 5:   # Credit — proper number format
                cell.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    v = float(str(cell.value).replace(",","")) if cell.value and str(cell.value).strip() not in ("","—","nan") else None
                except (ValueError, TypeError):
                    v = None
                if v:
                    cell.value         = v
                    cell.number_format = NUM_FMT
                    cell.fill = CRED_BG
                    cell.font = Font(name="Calibri", size=10, bold=True, color="2E7D32")
                else:
                    cell.value = None; cell.fill = bg
                    cell.font  = Font(name="Calibri", size=10, color="CCCCCC")
            elif c == 6:   # Bank
                cell.fill = PatternFill("solid", fgColor="EFF6FF")
                cell.font = Font(name="Calibri", size=9, bold=True, color="1D4ED8")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 7:   # Expenses Category
                cat_val = str(cell.value) if cell.value else ""
                bg2, fg2 = CAT_COLORS.get(cat_val, ("F9FAFB","374151"))
                cell.fill = PatternFill("solid", fgColor=bg2)
                cell.font = Font(name="Calibri", size=9, bold=True, color=fg2)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif c == 8:   # Expenses sub-category
                cell.fill = PatternFill("solid", fgColor="F0FDF4")
                cell.font = Font(name="Calibri", size=9, bold=True, color="166534")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 9:   # Income Category
                inc_val = str(cell.value) if cell.value else ""
                ib, ifg = INC_COLORS.get(inc_val, ("F9FAFB","374151"))
                cell.fill = PatternFill("solid", fgColor=ib)
                cell.font = Font(name="Calibri", size=9, bold=True, color=ifg)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif c == 10:  # Income sub-category
                cell.fill = INC_SUB
                cell.font = Font(name="Calibri", size=9, bold=True, color="065F46")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Data Validation Dropdowns ───────────────────────────────────────────
    data_range = f"2:{n_data + 1}"
    def make_dv(col_ref, show_err=True):
        dv = DataValidation(
            type       = "list",
            formula1   = f"=_Lists!${col_ref}$1:${col_ref}${max_len}",
            allow_blank= True,
            showErrorMessage = show_err,
        )
        dv.error       = "Please select a valid option from the list"
        dv.errorTitle  = "Invalid Entry"
        dv.prompt      = "Select from dropdown"
        dv.promptTitle = "Category"
        return dv

    last_row = n_data + 1
    dv_exp_cat = make_dv("A"); ws.add_data_validation(dv_exp_cat); dv_exp_cat.add(f"G2:G{last_row}")
    dv_inc_cat = make_dv("B"); ws.add_data_validation(dv_inc_cat); dv_inc_cat.add(f"I2:I{last_row}")
    dv_exp_sub = make_dv("C"); ws.add_data_validation(dv_exp_sub); dv_exp_sub.add(f"H2:H{last_row}")
    dv_inc_sub = make_dv("D"); ws.add_data_validation(dv_inc_sub); dv_inc_sub.add(f"J2:J{last_row}")

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Bank-to-Bank Transfers sheet ────────────────────────────────────────
    if transfer_df is not None and not transfer_df.empty:
        t_cols = ["Date","Description","Debit","Credit","Bank"]
        # Ensure Bank col (might still be _bank before rename at call site)
        if "Bank" not in transfer_df.columns and "_bank" in transfer_df.columns:
            transfer_df = transfer_df.rename(columns={"_bank": "Bank"})
        existing = [c for c in t_cols if c in transfer_df.columns]
        transfer_df[existing].to_excel(
            str(output_path), sheet_name="Bank Transfers",
            index=False, engine="openpyxl",
            startrow=0
        )
        # reload wb since to_excel with existing wb path can overwrite
        # → write to a temp then copy sheet
        import tempfile, shutil
        tmp = Path(str(output_path) + ".tmp.xlsx")
        transfer_df[existing].to_excel(str(tmp), sheet_name="Bank Transfers", index=False)
        wb_tmp = load_workbook(str(tmp))
        ws_tr  = wb_tmp["Bank Transfers"]
        ws_tr2 = wb.create_sheet("Bank Transfers")
        for row in ws_tr.iter_rows(values_only=True):
            ws_tr2.append(list(row))
        # Style transfer sheet header
        for cell in ws_tr2[1]:
            cell.fill = PatternFill("solid", fgColor="2D3748")
            cell.font = Font(name="Calibri", bold=True, color="F7FAFC", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i2, row2 in enumerate(ws_tr2.iter_rows(min_row=2), start=2):
            bg2 = PatternFill("solid", fgColor="FFF5F5") if i2%2==0 else PatternFill("solid", fgColor="FFFAFA")
            for cell in row2:
                cell.fill = bg2
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_l, w2 in zip("ABCDE",[13,60,14,14,16]):
            ws_tr2.column_dimensions[col_l].width = w2
        wb_tmp.close(); tmp.unlink()

    wb.save(str(output_path))

    tcr = round(df["_amt_cred"].fillna(0).sum(), 2)
    tdr = round(df["_amt_deb"].fillna(0).sum(),  2)
    return tcr, tdr, output_path


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 64)
    print("  Monthly Bank Statement Merger  v3")
    print(f"  Input  : {INPUT_FOLDER}")
    print(f"  Output : {OUTPUT_FOLDER}")
    print("=" * 64)

    # Collect all statement files in the same folder as the script
    exts  = {".csv", ".xls", ".xlsx"}
    # Ensure folders exist
    INPUT_FOLDER.mkdir(exist_ok=True)
    OUTPUT_FOLDER.mkdir(exist_ok=True)

    # Separate tradebook files from bank statement files
    all_input = sorted(
        p for p in INPUT_FOLDER.iterdir()
        if p.is_file()
        and p.suffix.lower() in exts
        and not p.name.startswith("~")
    )
    tradebook_files = [p for p in all_input if TRADEBOOK_PATTERN in p.stem.lower()]
    files           = [p for p in all_input if TRADEBOOK_PATTERN not in p.stem.lower()]

    if not files:
        print(f"\n[ERROR] No CSV / XLS / XLSX files found in:\n        {SCRIPT_DIR}")
        print("        Place your bank statement files there and re-run.")
        sys.exit(1)

    # Identify banks by filename
    print(f"\n  Found {len(files)} file(s) — detecting banks from filename…\n")
    identified = []
    for fp in files:
        display, parser_key = identify_bank(fp)
        marker = "✓" if parser_key != "Unknown" else "?"
        print(f"  {marker} [{display:<14}]  {fp.name}")
        identified.append((display, parser_key, fp))

    # Parse
    print("\n  Parsing transactions…\n")
    all_dfs = []
    for display, parser_key, fp in identified:
        parser = PARSERS.get(parser_key, _column_sniff_parse)
        try:
            df = parser(fp)
            if df.empty:
                print(f"  ✗ {display:<14}  0 rows  (check file format)  ←  {fp.name}")
                continue
            df["_bank"] = display          # full display name e.g. "Canara Bank"
            print(f"  ✓ {display:<14}  {len(df):>4} rows  ←  {fp.name}")
            all_dfs.append(df)
        except Exception as e:
            print(f"  ✗ {display:<14}  ERROR: {e}  ←  {fp.name}")

    if not all_dfs:
        print("\n[ERROR] No data could be parsed from any file. Aborting.")
        sys.exit(1)

    # Merge & compute helper columns
    merged = pd.concat(all_dfs, ignore_index=True)
    merged["_amt_deb"]   = merged["Debit"].apply(to_float)
    merged["_amt_cred"]  = merged["Credit"].apply(to_float)
    merged["_dt"]        = merged["Date"].apply(parse_date_str)
    merged["_norm_date"] = merged["_dt"].dt.strftime("%Y-%m-%d")

    print(f"\n  Total rows after merge  : {len(merged)}")

    # Remove bank-to-bank transfers (also captures them for the Transfers sheet)
    merged, transfer_df = remove_bank_transfers(merged)
    print(f"  Total rows after cleanup: {len(merged)}")

    # ── Zerodha: remove DEBIT rows (replaced by tradebook), keep CREDIT rows ──
    # Debit  = money sent to Zerodha for buying stocks → covered by tradebook detail
    # Credit = dividend / proceeds / refund from Zerodha → keep for income tracking
    is_zerodha     = merged["Description"].str.lower().str.contains("zerodha", na=False)
    is_zerodha_deb = is_zerodha & merged["_amt_deb"].notna() & (merged["_amt_deb"] > 0)
    is_zerodha_cr  = is_zerodha & merged["_amt_cred"].notna() & (merged["_amt_cred"] > 0)
    merged = merged[~is_zerodha_deb].copy()
    print(f"  Removed {is_zerodha_deb.sum()} Zerodha DEBIT rows (covered by tradebook)")
    if is_zerodha_cr.sum() > 0:
        print(f"  Kept    {is_zerodha_cr.sum()} Zerodha CREDIT rows (dividends / proceeds)")

    # ── Parse tradebook files & append (additive — all trades included) ──────
    trade_dfs = []
    if tradebook_files:
        print(f"  Tradebook file(s) found: {len(tradebook_files)}")
        for tp in tradebook_files:
            tdf = parse_tradebook(tp)
            if not tdf.empty:
                trade_dfs.append(tdf)
    else:
        print("  No tradebook file found in INPUT_FOLDER (optional)")

    if trade_dfs:
        trade_combined = pd.concat(trade_dfs, ignore_index=True)
        trade_combined["_dt"]        = trade_combined["Date"].apply(parse_date_str)
        trade_combined["_norm_date"] = trade_combined["_dt"].dt.strftime("%Y-%m-%d")
        merged = pd.concat([merged, trade_combined], ignore_index=True)
        print(f"  Added {len(trade_combined)} tradebook rows (all symbols, additive)")

    # Ensure sub-category columns exist and are clean (no 'nan' strings)
    for _sc in ["Expenses sub-category", "Income sub-category",
                "Expenses Category", "Income Category"]:
        if _sc not in merged.columns:
            merged[_sc] = ""
        merged[_sc] = merged[_sc].fillna("").astype(str).str.strip()
        merged[_sc] = merged[_sc].replace({"nan": "", "NaN": "", "None": ""})

    # Sort by date, renumber S.No
    merged.sort_values("_dt", inplace=True, na_position="last")
    merged.reset_index(drop=True, inplace=True)
    merged.rename(columns={"_bank": "Bank"}, inplace=True)
    merged["Bank"] = merged["Bank"].fillna("Zerodha")
    merged.insert(0, "S.No", range(1, len(merged) + 1))

    # ── Load category history from previous month files (Fix 4) ─────────────
    history = load_category_history(OUTPUT_FOLDER)

    # ── Determine which rows need expense categorization ──────────────────────
    # Rule: only DEBIT rows; credit-only rows stay blank in Expenses Category.
    is_debit       = merged["_amt_deb"].notna() & (merged["_amt_deb"] > 0)
    is_credit_only = (merged["_amt_cred"].notna() & (merged["_amt_cred"] > 0)) & ~is_debit
    has_no_cat     = merged["Expenses Category"].isna() | (merged["Expenses Category"] == "")
    needs_cat      = has_no_cat & ~is_credit_only

    # Pass 1 — History lookup (instant, zero API calls)
    history_hits = 0
    still_needs  = []
    for idx in merged.index[needs_cat]:
        desc  = str(merged.at[idx, "Description"])
        match = match_history(desc, history)
        if match:
            cat, sub = match
            merged.at[idx, "Expenses Category"]     = cat
            if sub and not merged.at[idx, "Expenses sub-category"]:
                merged.at[idx, "Expenses sub-category"] = sub
            history_hits += 1
        else:
            still_needs.append(idx)
    print(f"  History matched : {history_hits} rows")

    # Pass 2 — Keyword rules + AI for remaining rows
    print(f"\n{'─'*64}")
    print(f"  AI Categorizing {len(still_needs)} remaining transactions via NVIDIA API…")
    print(f"{'─'*64}")

    if still_needs:
        descs      = merged.loc[still_needs, "Description"].tolist()
        cats, subs = categorize_transactions(descs)
        for idx, cat, sub in zip(still_needs, cats, subs):
            merged.at[idx, "Expenses Category"] = cat
            if sub and not merged.at[idx, "Expenses sub-category"]:
                merged.at[idx, "Expenses sub-category"] = sub
    else:
        print("    All rows matched from history or pre-filled — no API calls needed")
    print(f"  ✓ Expense categorization complete")

    # ── Income Category + sub-category (credit rows only) ────────────────────
    print(f"  Categorizing Income (credit rows) via hybrid rules…")
    if "Income Category" not in merged.columns:
        merged["Income Category"]     = ""
        merged["Income sub-category"] = ""

    credit_mask = merged["_amt_cred"].notna() & (merged["_amt_cred"] > 0)
    for idx in merged.index[credit_mask]:
        desc = str(merged.at[idx, "Description"])
        cat, sub = apply_income_rules(desc)
        if not merged.at[idx, "Income Category"]:
            merged.at[idx, "Income Category"]     = cat
            merged.at[idx, "Income sub-category"] = sub
    print(f"  ✓ Income categorization complete  ({credit_mask.sum()} credit rows tagged)")

    # ── Compute output filename from last transaction date (Fix 3) ──────────
    valid_dates = merged["_dt"].dropna()
    last_date   = valid_dates.max().strftime("%Y-%m-%d") if not valid_dates.empty else datetime.now().strftime("%Y-%m-%d")
    output_file = OUTPUT_FOLDER / f"Full_statement_{last_date}.xlsx"
    print(f"  Output file : Full_statement_{last_date}.xlsx")

    # Write Excel
    tcr, tdr, saved_path = write_excel(merged, transfer_df, output_file)

    print(f"\n{'═'*64}")
    print(f"  ✅  Saved → {saved_path}")
    print(f"{'─'*64}")
    print(f"  Total Credit : ₹{tcr:>12,.2f}")
    print(f"  Total Debit  : ₹{tdr:>12,.2f}")
    print(f"  Net          : ₹{tcr - tdr:>12,.2f}")
    print(f"{'═'*64}")


if __name__ == "__main__":
    main()
